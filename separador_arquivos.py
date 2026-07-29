from openpyxl import load_workbook, Workbook
from pathlib import Path

# ==========================================
# CONFIGURAÇÕES
# ==========================================

ARQUIVO_ENTRADA = "dados_medicos_base_v1.xlsx"

PASTA_SAIDA = "xlsx_dividido"

# Quantidade de linhas por arquivo
# (não conta a linha do cabeçalho)
LINHAS_POR_ARQUIVO = 100000

# ==========================================

Path(PASTA_SAIDA).mkdir(exist_ok=True)

print("Abrindo arquivo...")

workbook = load_workbook(ARQUIVO_ENTRADA, read_only=True)
worksheet = workbook.active

# Lê o cabeçalho
cabecalho = next(worksheet.iter_rows(values_only=True))

contador_arquivo = 1
contador_linhas = 0

novo_wb = Workbook()
nova_ws = novo_wb.active
nova_ws.title = worksheet.title

# Escreve o cabeçalho
nova_ws.append(cabecalho)

for linha in worksheet.iter_rows(min_row=2, values_only=True):

    nova_ws.append(linha)
    contador_linhas += 1

    if contador_linhas >= LINHAS_POR_ARQUIVO:

        nome = Path(PASTA_SAIDA) / f"dataset_medico_part{contador_arquivo:03}.xlsx"

        novo_wb.save(nome)

        print(f"Criado: {nome}")

        contador_arquivo += 1
        contador_linhas = 0

        novo_wb = Workbook()
        nova_ws = novo_wb.active
        nova_ws.title = worksheet.title
        nova_ws.append(cabecalho)

# Salva a última parte (caso ainda haja linhas)
if contador_linhas > 0:

    nome = Path(PASTA_SAIDA) / f"dataset_medico_part{contador_arquivo:03}.xlsx"

    novo_wb.save(nome)

    print(f"Criado: {nome}")

workbook.close()

print("Processo concluído!")